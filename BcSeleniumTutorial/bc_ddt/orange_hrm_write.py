'''
Created on 23 Jun 2026

@author: Bharath.c
'''
from openpyxl import load_workbook

#1 loading the workbook
filename = r"C:\Users\Bharath.c\Documents\testdata\orangehrmtestdata.xlsx"
my_workbook = load_workbook(filename)

#2 set the active page

Sheet1 = my_workbook["Sheet1"]

total_cols = Sheet1.max_column
print("total_cols:",total_cols)

total_rows = Sheet1.max_row
print("total_rows:",total_rows)
print("=====================")

for i in range(2,total_rows+1):
    Sheet1.cell(i,5).value = "passed"
my_workbook.save(filename)