'''
Created on 22 Jun 2026

@author: Bharath.c
'''
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#1 loading the workbook
filename = r"C:\Users\Bharath.c\Documents\testdata\orangehrmtestdata.xlsx"
my_workbook = load_workbook(filename)

#2 set the active page

Sheet1 = my_workbook["Sheet1"]

total_cols = Sheet1.max_column
print("total_cols:",total_cols)

total_rows = Sheet1.max_row
print("total_rows:",total_rows)
print("=====================")

for i in range(2,total_rows+1):
    #3 featch and print cell value
    username = Sheet1.cell(i,2).value
    print("username:",username)
    
    password = Sheet1.cell(i,3).value
    print("password :",password)
    
    print("=====================")

    # 1. Launch Chrome browser
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)
    options.add_argument("start-maximized")
        
    driver = webdriver.Chrome(options)
    driver.implicitly_wait(10)
        
    # 2. Navigate to OrangeHRM Login page
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    
    wait = WebDriverWait(driver,timeout=10)
        
    # 3. Enter Username
    user_wrt = wait.until(EC.visibility_of_element_located((By.NAME, 'username')))
    #user_wrt = driver.find_element(By.NAME, 'username')
    user_wrt.send_keys(username)
        
    # 4. Enter Password
    pass_word = wait.until(EC.visibility_of_element_located((By.NAME, 'password')))
    #pass_word = driver.find_element(By.NAME, 'password')
    pass_word.send_keys(password)
        
    # 5. Click on Login Button
    login_clk = driver.find_element(By.TAG_NAME,'button')
    login_clk.click()

    #6 validate weather dashboards is present in current page url
    current_page_url = driver.current_url
    print(driver.current_url)
    expected_url = Sheet1.cell(i,4).value
    print("expected_url :",expected_url)
    if expected_url in current_page_url:
        print("Test case passed")
        Sheet1.cell(i,5).value = "passed"
        my_workbook.save(filename)
    else:
        print("Test case failed")
        Sheet1.cell(i,5).value = "failed"
        my_workbook.save(filename)
    print("=====================")