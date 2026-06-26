'''
Created on 18 Jun 2026

@author: Bharath.c
'''
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

#1 loading the workbook
filename = r"C:\Users\Bharath.c\Documents\testdata\orangehrmtestdata.xlsx"
my_workbook = load_workbook(filename)

#2 set the active page

Sheet1 = my_workbook["Sheet1"]

for i in range(2,6):
    #3 featch and print cell value
    username = Sheet1.cell(i,2).value
    print("username:",username)
    
    password = Sheet1.cell(i,3).value
    print("password :",password)

    # 1. Launch Chrome browser
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)
    options.add_argument("start-maximized")
    
    driver = webdriver.Chrome(options)
    driver.implicitly_wait(10)
    
    # 2. Navigate to OrangeHRM Login page
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    
    # 3. Enter Username
    user_wrt = driver.find_element(By.NAME, 'username')
    user_wrt.send_keys(username)
    
    # 4. Enter Password
    pass_word = driver.find_element(By.NAME, 'password')
    pass_word.send_keys(password)
    
    # 5. Click on Login Button
    login_clk = driver.find_element(By.TAG_NAME,'button')
    login_clk.click()